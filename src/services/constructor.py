# -*- coding: utf-8 -*-
"""
Сервис конструктора сводных таблиц (Low-code).
Позволяет загрузить Excel-файл, просмотреть данные,
применить фильтры, построить сводную таблицу и выгрузить результат.
"""

import os
import re
import json
import uuid
import traceback
import logging
import warnings
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

import pandas as pd
import numpy as np

from ..types import PivotResult, ApplyFiltersResult

logger = logging.getLogger(__name__)

# Константы
AGG_COLUMN_NAME = '__agg__'
"""Имя колонки, указывающей тип агрегации в иерархическом формате."""

DATE_PREFIXES = {'__год__', '__квартал__', '__месяц__', '__день__'}
"""Префиксы для виртуальных колонок декомпозиции дат."""

# Директория для хранения сценариев
SCENARIOS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config', 'constructor_scenarios')


def _ensure_scenarios_dir():
    """Создаёт директорию для сценариев, если её нет."""
    if not os.path.exists(SCENARIOS_DIR):
        os.makedirs(SCENARIOS_DIR, exist_ok=True)


def _detect_header_row(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """
    Автоматически определяет, в какой строке находятся заголовки,
    и возвращает информацию для пользовательского выбора.

    Алгоритм:
    1. Читает первые N строк с header=None
    2. Для каждой строки считает количество не-NaN значений
    3. Пробует header=0, header=1, header=2 и считает unnamed-колонки
    4. Возвращает best_guess + все строки для предпросмотра

    :param file_path: Путь к Excel-файлу
    :param sheet_name: Имя листа
    :return: {best_header_row, unnamed_ratio, rows_preview: [{row_index, values, n_non_null}, ...]}
    """
    # Читаем первые 5 строк без заголовков
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=6)

    rows_preview = []
    best_row = 0
    best_unnamed_ratio = 1.0

    for i in range(min(6, len(df_raw))):
        row_values = df_raw.iloc[i].tolist()
        non_null = sum(1 for v in row_values if pd.notna(v))
        rows_preview.append({
            'row_index': i,
            'values': [str(v) if pd.notna(v) else '' for v in row_values],
            'n_non_null': non_null,
            'total_cols': len(row_values),
        })

    # Пробуем разные header_row и считаем unnamed
    for header_row in range(min(4, len(df_raw))):
        try:
            df_test = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=5)
            cols = df_test.columns.tolist()
            unnamed_count = sum(1 for c in cols if 'Unnamed' in str(c))
            total = len(cols)
            ratio = unnamed_count / total if total > 0 else 1.0
            if ratio < best_unnamed_ratio:
                best_unnamed_ratio = ratio
                best_row = header_row
        except Exception:
            pass

    # Если best_unnamed_ratio всё ещё > 0.5 — скорее всего нет заголовков
    # или структура сложная. Оставляем best_row=0 как значение по умолчанию.

    return {
        'best_header_row': best_row,
        'unnamed_ratio': best_unnamed_ratio,
        'rows_preview': rows_preview,
        'needs_review': best_unnamed_ratio > 0.3,  # >30% unnamed — нужна проверка
    }


def load_excel_file(file_path: str) -> Dict[str, Any]:
    """
    Загружает Excel-файл и возвращает список листов с метаданными.

    :param file_path: Путь к файлу .xlsx/.xls
    :return: {sheets: [{name, rows, cols, suggested_header_row, needs_review}, ...]}
    """
    xls = pd.ExcelFile(file_path)
    sheets = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
        # Пробуем прочитать пару строк для оценки
        df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
        # Автоопределение заголовков
        header_info = _detect_header_row(file_path, sheet_name)
        sheets.append({
            'name': sheet_name,
            'rows': len(df_sample) + 100,  # приблизительно
            'cols': len(df.columns),
            'columns': df.columns.tolist(),
            'suggested_header_row': header_info['best_header_row'],
            'needs_review': header_info['needs_review'],
        })
    return {'sheets': sheets}


def load_sheet_data(
    file_path: str,
    sheet_name: str,
    limit: int = 100,
    offset: int = 0,
    header_row: Optional[int] = None,
    transpose: bool = False,
) -> Dict[str, Any]:
    """
    Загружает данные указанного листа Excel-файла.

    :param file_path: Путь к файлу
    :param sheet_name: Имя листа
    :param limit: Макс. количество строк
    :param offset: Смещение
    :param header_row: Номер строки с заголовками (0-based). None = автоопределение.
    :param transpose: Если True, транспонирует данные (строки -> колонки, колонки -> строки).
                     Первая строка после транспонирования становится заголовками.
    :return: {columns, data, total_rows, dtypes, date_columns, header_row_used}
    """
    # Автоопределение строки заголовков, если не указана
    if header_row is None:
        header_info = _detect_header_row(file_path, sheet_name)
        header_row = header_info['best_header_row']

    if transpose:
        # При транспонировании: читаем без заголовков, затем транспонируем
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str)
        df_raw = df_raw.fillna('').astype(str)
        df_t = df_raw.T.reset_index(drop=True)
        # Первая строка транспонированных данных становится заголовками
        if len(df_t) > 0:
            new_columns = df_t.iloc[0].tolist()
            # Переименовываем Unnamed-колонки в понятные имена
            new_columns = [str(c) if pd.notna(c) and str(c).strip() else f'Column_{i}'
                          for i, c in enumerate(new_columns)]
            df_t.columns = new_columns
            df_t = df_t.iloc[1:].reset_index(drop=True)
        columns = df_t.columns.tolist()
        df = df_t
    else:
        # Читаем meta с указанной строкой заголовков
        df_meta = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=0)
        columns = df_meta.columns.tolist()

        # Читаем данные с указанной строкой заголовков
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, dtype=str)

        # Заменяем NaN на пустую строку и приводим всё к строке
        df = df.fillna('').astype(str)

    total_rows = len(df)

    # Применяем offset и limit
    if offset > 0:
        df = df.iloc[offset:]
    df = df.head(limit)

    data = df.to_dict(orient='records')

    # Определяем типы колонок (по первым 1000 строкам)
    if transpose:
        # При транспонировании определяем типы по самим данным (уже загружены)
        dtypes = _infer_column_types_from_df(df, columns)
    else:
        dtypes = _infer_column_types(file_path, sheet_name, columns, header_row=header_row)

    # Определяем, какие колонки являются датами
    date_columns = [col for col, dtype in dtypes.items() if dtype == 'date']

    return {
        'columns': columns,
        'data': data,
        'total_rows': total_rows,
        'dtypes': dtypes,
        'date_columns': date_columns,
        'header_row_used': header_row if not transpose else None,
    }


def _parse_dates_flexible(series: pd.Series) -> pd.Series:
    """
    Универсальный парсер дат, поддерживающий:
    - ISO-формат: 2026-05-01, 2026-05-01 15:12:16
    - Русский формат: 01.05.2026, 31.12.2026, 01/05/2026

    Алгоритм:
    1. По умолчанию парсим с dayfirst=True (для русского ДД.ММ.ГГГГ)
    2. Для значений, начинающихся с 4 цифр (ISO-формат 2026-...),
       перепарсиваем без dayfirst, чтобы избежать путаницы день/месяц

    format='mixed' необходим для pandas 3.x, где to_datetime без указания
    формата не справляется со смешанными форматами (с микросекундами и без).
    """
    series_str = series.astype(str).str.strip()
    # Шаг 1: dayfirst=True для русских дат (ДД.ММ.ГГГГ)
    # Подавляем warning "Parsing dates in %Y-%m-%d format when dayfirst=True"
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', UserWarning)
        result = pd.to_datetime(series_str, errors='coerce', dayfirst=True, format='mixed')
    # Шаг 2: перепарсиваем ISO-даты (начинаются с 4 цифр — год)
    # без dayfirst, т.к. 2026-05-01 с dayfirst=True даёт день=5, месяц=1
    iso_mask = series_str.str.match(r'^\d{4}') & series_str.notna() & (series_str != '')
    if iso_mask.any():
        result[iso_mask] = pd.to_datetime(series_str[iso_mask], errors='coerce', format='mixed')
    return result


def _infer_column_types(
    file_path: str,
    sheet_name: str,
    columns: List[str],
    header_row: int = 0,
) -> Dict[str, str]:
    """
    Определяет типы колонок (number, date, text) по данным.
    
    Приоритет: number → date → text.
    Дата определяется ТОЛЬКО если колонка не является числовой,
    чтобы числа (в т.ч. Excel serial dates) не помечались как даты.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=1000, dtype=str)
    except Exception:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=1000, dtype=str)

    # Регулярка для определения "похожести на дату":
    # значение содержит точки, слеши или дефисы между цифрами
    # например: 01.04.2026, 2026-04-01, 01/04/26, 01.04.2026 15:12
    date_pattern = re.compile(
        r'\d{1,4}[-./]\d{1,2}[-./]\d{1,4}'  # дата с разделителями
    )

    dtypes = {}
    for col in columns:
        if col not in df.columns:
            dtypes[col] = 'text'
            continue

        series = df[col].dropna()
        if len(series) == 0:
            dtypes[col] = 'text'
            continue

        # 1. Проверка на ЧИСЛО (приоритет — перед датой)
        series_str = series.astype(str)
        try:
            numeric_count = pd.to_numeric(series_str, errors='coerce').notna().sum()
            if numeric_count > len(series) * 0.5:
                dtypes[col] = 'number'
                continue
        except Exception:
            pass

        # 2. Проверка на ДАТУ (только если колонка НЕ числовая)
        # Дополнительно проверяем, что значения выглядят как даты
        try:
            looks_like_date = series_str.str.contains(date_pattern, na=False).sum()
        except Exception:
            looks_like_date = 0
        if looks_like_date > len(series) * 0.3:
            try:
                # dayfirst=True НЕ используем — Excel даты приходят в ISO-формате (2026-05-01),
                # где dayfirst=True ломает парсинг (05 → день, 01 → месяц)
                date_count = _parse_dates_flexible(series_str).notna().sum()
                if date_count > len(series) * 0.5:
                    dtypes[col] = 'date'
                    continue
            except Exception:
                pass

        # 3. Иначе — текст
        dtypes[col] = 'text'

    return dtypes


def _infer_column_types_from_df(df: pd.DataFrame, columns: List[str]) -> Dict[str, str]:
    """
    Определяет типы колонок по уже загруженному DataFrame (без перечитывания файла).
    Используется при транспонировании данных.

    :param df: DataFrame со строковыми данными
    :param columns: Список колонок
    :return: {col_name: 'number'|'date'|'text', ...}
    """
    date_pattern = re.compile(
        r'\d{1,4}[-./]\d{1,2}[-./]\d{1,4}'
    )
    dtypes = {}
    for col in columns:
        if col not in df.columns:
            dtypes[col] = 'text'
            continue
        series = df[col].dropna().astype(str)
        if len(series) == 0:
            dtypes[col] = 'text'
            continue
        # 1. Проверка на ЧИСЛО
        try:
            numeric_count = pd.to_numeric(series, errors='coerce').notna().sum()
            if numeric_count > len(series) * 0.5:
                dtypes[col] = 'number'
                continue
        except Exception:
            pass
        # 2. Проверка на ДАТУ
        try:
            looks_like_date = series.str.contains(date_pattern, na=False).sum()
        except Exception:
            looks_like_date = 0
        if looks_like_date > len(series) * 0.3:
            try:
                date_count = _parse_dates_flexible(series).notna().sum()
                if date_count > len(series) * 0.5:
                    dtypes[col] = 'date'
                    continue
            except Exception:
                pass
        # 3. Текст
        dtypes[col] = 'text'
    return dtypes


def decompose_date_column(series: pd.Series, col_name: str) -> Dict[str, pd.Series]:
    """
    Разбивает datetime-колонку на компоненты: год, квартал, месяц, день.
    Месяц возвращается названием (Январь, Февраль...).

    :param series: Исходная серия с датами (строки)
    :param col_name: Имя колонки для формирования ключей
    :return: {__год__col: Series, __квартал__col: Series, ...}
    """
    # Приводим к строке перед преобразованием в datetime, чтобы избежать mixed types
    # dayfirst=True НЕ используем — Excel даты приходят в ISO-формате (2026-05-01),
    # где dayfirst=True ломает парсинг (05 → день, 01 → месяц).
    # ISO-формат: 2026-05-01 → без dayfirst: day=1, month=5 ✓
    #            с  dayfirst: day=5, month=1 ✗
    dt_series = _parse_dates_flexible(series)
    nat_mask = dt_series.isna()

    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    # Сначала вычисляем значения через fillna(0) для корректной типизации,
    # затем заменяем строки, где исходная дата была NaT, на пустую строку
    year_vals = dt_series.dt.year.fillna(0).astype(int).astype(str)
    year_vals[nat_mask] = ''

    quarter_vals = 'Q' + dt_series.dt.quarter.fillna(0).astype(int).astype(str)
    quarter_vals[nat_mask] = ''

    month_vals = dt_series.dt.month.map(month_names).fillna('')

    day_vals = dt_series.dt.day.fillna(0).astype(int).astype(str).str.zfill(2)
    day_vals[nat_mask] = ''

    return {
        f'__год__{col_name}': year_vals,
        f'__квартал__{col_name}': quarter_vals,
        f'__месяц__{col_name}': month_vals,
        f'__день__{col_name}': day_vals,
    }


def _apply_date_decomposition(df: pd.DataFrame, row_cols: List[str]) -> pd.DataFrame:
    """
    Проверяет row_cols на префиксы __год__, __квартал__, __месяц__, __день__.
    Если найдены — разбивает datetime-колонку на компоненты и добавляет в df.

    :param df: Исходный DataFrame
    :param row_cols: Список колонок (могут содержать __год__Имя, __месяц__Имя и т.д.)
    :return: DataFrame с добавленными виртуальными колонками
    """
    # Собираем уникальные родительские колонки, которые нужно разложить
    parent_cols = set()
    for col in row_cols:
        for prefix in DATE_PREFIXES:
            if col.startswith(prefix):
                parent = col[len(prefix):]
                if parent in df.columns:
                    parent_cols.add(parent)

    for parent in parent_cols:
        decomposed = decompose_date_column(df[parent], parent)
        for key, series in decomposed.items():
            df[key] = series.values

    return df


def _apply_filters_to_df(
    df: pd.DataFrame,
    filters: Optional[Dict[str, Dict[str, Any]]],
) -> pd.DataFrame:
    """
    Применяет фильтры к DataFrame (общая функция для apply_filters и build_pivot_table).

    Поддерживаемые типы фильтров:
    - equals / not_equals — точное совпадение (регистронезависимое)
    - contains — поиск подстроки
    - greater_than / less_than — числовое сравнение
    - is_empty / is_not_empty — проверка на пустоту

    :param df: Исходный DataFrame
    :param filters: {column: {type, value}}
    :return: Отфильтрованный DataFrame
    """
    if not filters:
        return df

    for col, filter_def in filters.items():
        if col not in df.columns:
            continue
        filter_type = filter_def.get('type', '')
        filter_value = filter_def.get('value', '')

        if filter_type == 'equals' and filter_value != '':
            df = df[df[col].astype(str).str.lower() == str(filter_value).lower()]
        elif filter_type == 'contains' and filter_value != '':
            df = df[df[col].astype(str).str.lower().str.contains(str(filter_value).lower(), na=False)]
        elif filter_type == 'not_equals' and filter_value != '':
            df = df[df[col].astype(str).str.lower() != str(filter_value).lower()]
        elif filter_type == 'greater_than':
            try:
                num_val = float(filter_value)
                tmp_col = col + '_num'
                df[tmp_col] = pd.to_numeric(df[col], errors='coerce')
                df = df[df[tmp_col] > num_val]
                df = df.drop(columns=[tmp_col])
            except (ValueError, TypeError):
                pass
        elif filter_type == 'less_than':
            try:
                num_val = float(filter_value)
                tmp_col = col + '_num'
                df[tmp_col] = pd.to_numeric(df[col], errors='coerce')
                df = df[df[tmp_col] < num_val]
                df = df.drop(columns=[tmp_col])
            except (ValueError, TypeError):
                pass
        elif filter_type == 'is_empty':
            df = df[df[col].astype(str) == '']
        elif filter_type == 'is_not_empty':
            df = df[df[col].astype(str) != '']

    return df


def apply_filters(
    file_path: str,
    sheet_name: str,
    selected_columns: Optional[List[str]] = None,
    filters: Optional[Dict[str, Dict[str, Any]]] = None,
    sort_column: Optional[str] = None,
    sort_order: str = 'asc',
    limit: int = 100,
    offset: int = 0,
    cached_df: Optional[pd.DataFrame] = None,
    header_row: int = 0,
) -> Dict[str, Any]:
    """
    Применяет фильтры и возвращает отфильтрованные данные.

    :param file_path: Путь к файлу
    :param sheet_name: Имя листа
    :param selected_columns: Список колонок для отображения (None = все)
    :param filters: Словарь фильтров {column: {type, value}}
    :param sort_column: Колонка для сортировки
    :param sort_order: 'asc' или 'desc'
    :param limit: Макс. строк
    :param offset: Смещение
    :param header_row: Номер строки с заголовками (0-based)
    :return: {columns, data, total_rows, filtered_count}
    """
    # Читаем все данные (используем кэш, если передан)
    if cached_df is not None:
        df = cached_df.copy().astype(str)
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, dtype=str)
    # Приводим всё к строке, чтобы избежать datetime объектов при JSON-сериализации
    df = df.fillna('').astype(str)

    total_rows = len(df)

    # Выбор колонок
    if selected_columns and len(selected_columns) > 0:
        valid_cols = [c for c in selected_columns if c in df.columns]
        if valid_cols:
            df = df[valid_cols]

    # Применение фильтров (общая функция)
    df = _apply_filters_to_df(df, filters)

    filtered_count = len(df)

    # Сортировка
    if sort_column and sort_column in df.columns:
        ascending = sort_order == 'asc'
        df = df.sort_values(by=sort_column, ascending=ascending)

    # Пагинация
    if offset > 0:
        df = df.iloc[offset:]
    df = df.head(limit)

    data = df.to_dict(orient='records')
    columns = df.columns.tolist()

    return {
        'columns': columns,
        'data': data,
        'total_rows': total_rows,
        'filtered_count': filtered_count,
    }


def build_pivot_table(
    file_path: str,
    sheet_name: str,
    rows: List[str],
    values: List[str],
    cols: Optional[List[str]] = None,
    agg_functions: Optional[List[str]] = None,
    filters: Optional[Dict[str, Dict[str, Any]]] = None,
    output_format: str = 'flat',
    cached_df: Optional[pd.DataFrame] = None,
    totals_mode: str = 'none',
    header_row: int = 0,
) -> Dict[str, Any]:
    """
    Строит сводную таблицу на основе данных Excel-файла.
    Поддерживает декомпозицию дат (колонки с префиксами __год__, __месяц__ и т.д.),
    режим "Без изменений" (agg_functions=['none']),
    множественные функции агрегации и два формата вывода.

    :param file_path: Путь к файлу
    :param sheet_name: Имя листа
    :param rows: Колонки для строк сводной таблицы
    :param values: Колонки для значений
    :param cols: Колонки для столбцов сводной таблицы (опционально)
    :param agg_functions: Список функций агрегации (sum, mean, count, min, max, nunique, none)
    :param filters: Дополнительные фильтры
    :param output_format: Формат вывода: 'flat' (широкий) или 'hierarchical' (вертикальный)
    :param cached_df: Кэшированный DataFrame
    :param totals_mode: Режим итогов: 'none' — без итогов, 'rows' — итоговая колонка (по строкам),
                        'cols' — итоговая строка (по столбцам), 'both' — и строка, и колонка
    :param header_row: Номер строки с заголовками (0-based)
    :return: {pivot_data, columns, total_rows, aggregation(s), output_format}
    """
    if agg_functions is None:
        agg_functions = ['sum']

    # Читаем данные (используем кэш, если передан)
    if cached_df is not None:
        df = cached_df.copy().astype(str)
    else:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, dtype=str)
    # Приводим всё к строке, чтобы избежать datetime объектов при JSON-сериализации
    df = df.fillna('').astype(str)

    # Применяем фильтры если есть (общая функция)
    df = _apply_filters_to_df(df, filters)

    # Применяем декомпозицию дат к rows и cols
    all_dim_cols = list(rows) + (cols or [])
    df = _apply_date_decomposition(df, all_dim_cols)

    # Очищаем виртуальные префиксы из списков для работы с реальными колонками df
    def resolve_cols(col_list):
        """Возвращает колонки, которые реально есть в df (в т.ч. виртуальные)."""
        result = []
        for c in col_list:
            if c in df.columns:
                result.append(c)
        return result

    valid_rows = resolve_cols(rows)
    valid_values = resolve_cols(values)
    valid_cols = resolve_cols(cols or [])

    # Фильтруем строки с пустыми значениями в декомпозированных дата-колонках
    # (нераспарсенные даты не должны создавать пустые группы в сводной таблице)
    for row_col in valid_rows:
        for prefix in DATE_PREFIXES:
            if row_col.startswith(prefix):
                df = df[df[row_col].str.strip() != '']
                break

    if not valid_rows or not valid_values:
        return {
            'error': 'Не указаны строки или значения для сводной таблицы',
            'pivot_data': [],
            'columns': [],
            'total_rows': 0,
            'aggregations': agg_functions,
            'output_format': output_format,
        }

    # Режим "Без изменений" — возвращаем данные как есть
    if 'none' in agg_functions:
        # Выбираем только нужные колонки
        display_cols = valid_rows + valid_values + valid_cols
        # Убираем дубликаты, сохраняя порядок
        seen = set()
        display_cols_uniq = []
        for c in display_cols:
            if c not in seen:
                seen.add(c)
                display_cols_uniq.append(c)
        display_cols = display_cols_uniq

        result_df = df[display_cols] if all(c in df.columns for c in display_cols) else df

        total_rows = len(result_df)
        pivot_data = result_df.to_dict(orient='records')
        pivot_columns = result_df.columns.tolist()

        return {
            'pivot_data': pivot_data,
            'columns': pivot_columns,
            'total_rows': total_rows,
            'aggregations': ['none'],
            'output_format': output_format,
            'message': 'Данные без агрегации',
        }

    # Преобразуем значения в числа для агрегации (сохраняем NaN для пустых ячеек)
    for val_col in valid_values:
        df[val_col] = pd.to_numeric(df[val_col], errors='coerce')

    # Маппинг функций агрегации
    agg_map = {
        'sum': 'sum',
        'mean': 'mean',
        'count': 'count',
        'min': 'min',
        'max': 'max',
        'nunique': pd.Series.nunique,
    }
    agg_labels = {
        'sum': 'Сумма',
        'mean': 'Среднее',
        'count': 'Количество',
        'min': 'Минимум',
        'max': 'Максимум',
        'nunique': 'Уникальных',
    }

    try:
        # Вычисляем агрегацию для каждой функции отдельно
        agg_results = []  # список DataFrame'ов с результатами
        for agg_func_key in agg_functions:
            agg_func = agg_map.get(agg_func_key, 'sum')
            agg_label = agg_labels.get(agg_func_key, agg_func_key)

            # Для count-агрегации заменяем 0 и NaN на NaN, чтобы:
            # - пустые ячейки (где значения нет) не учитывались
            # - нулевые значения (чаевые = 0, скидка = 0) не учитывались
            # count должен считать только "значимые" (ненулевые) значения.
            # Для остальных агрегаций (sum, mean, min, max) NaN → 0.
            if agg_func_key != 'count':
                agg_df = df.copy()
                for val_col in valid_values:
                    agg_df[val_col] = agg_df[val_col].fillna(0)
            else:
                agg_df = df.copy()
                for val_col in valid_values:
                    # float dtype после pd.to_numeric, поэтому используем np.nan
                    agg_df[val_col] = agg_df[val_col].replace(0, np.nan)

            if valid_cols:
                # Сводная с колонками
                use_margins = totals_mode in ('rows', 'cols', 'both')
                pivot = pd.pivot_table(
                    agg_df,
                    values=valid_values,
                    index=valid_rows,
                    columns=valid_cols,
                    aggfunc=agg_func,
                    fill_value=0,
                    margins=use_margins,
                    margins_name='Итого',
                )
                pivot = pivot.reset_index()
                # Плоские колонки
                if isinstance(pivot.columns, pd.MultiIndex):
                    flat_cols = []
                    for col in pivot.columns:
                        if isinstance(col, tuple):
                            parts = [str(c) for c in col if str(c) != '']
                            # Первая часть — имя value колонки, остальные — значения col dimension
                            if len(parts) > 1:
                                if output_format == 'hierarchical':
                                    # Иерархический: колонка __agg__ покажет тип агрегации, суффикс не нужен
                                    flat_cols.append(f"{parts[0]} | {' | '.join(parts[1:])}")
                                else:
                                    # Плоский: суффикс агрегации в имени колонки
                                    flat_cols.append(f"{parts[0]} ({agg_label}) | {' | '.join(parts[1:])}")
                            else:
                                # row-колонка (второй элемент тупла пустой) — оставляем как есть
                                flat_cols.append(str(parts[0]))
                        else:
                            flat_cols.append(str(col))
                    pivot.columns = flat_cols

                    # Фильтруем "Итого" колонки и строки в зависимости от режима
                    if use_margins:
                        # Определяем "Итого" колонки (заканчиваются на "| Итого")
                        итого_cols_mask = [c.endswith(' | Итого') for c in flat_cols]
                        # Определяем "Итого" строки (значение "Итого" в любой row-колонке)
                        итого_rows_mask = pd.Series(False, index=pivot.index)
                        for rc in valid_rows:
                            if rc in pivot.columns:
                                итого_rows_mask = итого_rows_mask | (pivot[rc].astype(str).str.strip() == 'Итого')

                        if totals_mode == 'none':
                            # Удаляем и колонки, и строки
                            pivot = pivot.loc[~итого_rows_mask]
                            # удаляем колонки
                            cols_to_drop = [flat_cols[i] for i, is_итого in enumerate(итого_cols_mask) if is_итого]
                            if cols_to_drop:
                                pivot = pivot.drop(columns=cols_to_drop)
                                flat_cols = [c for c in flat_cols if c not in cols_to_drop]
                        elif totals_mode == 'rows':
                            # Только итоговая колонка (по строкам) — удаляем "Итого" строки
                            pivot = pivot.loc[~итого_rows_mask]
                        elif totals_mode == 'cols':
                            # Только итоговая строка (по столбцам) — удаляем "Итого" колонки
                            cols_to_drop = [flat_cols[i] for i, is_итого in enumerate(итого_cols_mask) if is_итого]
                            if cols_to_drop:
                                pivot = pivot.drop(columns=cols_to_drop)
                                flat_cols = [c for c in flat_cols if c not in cols_to_drop]
                        # totals_mode == 'both': ничего не удаляем, оставляем как есть
                else:
                    if output_format == 'hierarchical':
                        # Иерархический: не добавляем суффикс агрегации к value-колонкам
                        pass
                    else:
                        # Плоский: добавляем суффикс агрегации
                        rename_map = {}
                        for c in pivot.columns:
                            if c not in valid_rows:
                                rename_map[c] = f"{c} ({agg_label})"
                        if rename_map:
                            pivot = pivot.rename(columns=rename_map)
            else:
                # Простая группировка без колонок
                group_cols = valid_rows
                pivot = agg_df.groupby(group_cols, as_index=False)[valid_values].agg(agg_func)
                if isinstance(pivot.columns, pd.MultiIndex):
                    pivot.columns = [' '.join(col).strip() for col in pivot.columns.values]

                if output_format == 'hierarchical':
                    # Иерархический: не добавляем суффикс агрегации к value-колонкам
                    pass
                else:
                    # Плоский: добавляем суффикс агрегации
                    rename_map = {}
                    for c in pivot.columns:
                        if c not in valid_rows:
                            rename_map[c] = f"{c} ({agg_label})"
                    if rename_map:
                        pivot = pivot.rename(columns=rename_map)

                # Итоги для группировки без колонок
                if totals_mode in ('cols', 'both'):
                    # Итоговая строка (по столбцам) — сумма каждой колонки значений
                    total_vals = {rc: 'Итого' for rc in valid_rows}
                    for val_col in valid_values:
                        if output_format == 'hierarchical':
                            col_name = val_col
                        else:
                            col_name = f"{val_col} ({agg_label})"
                        total_vals[col_name] = round(float(agg_df[val_col].agg(agg_func)), 2)
                    pivot = pd.concat([pivot, pd.DataFrame([total_vals])], ignore_index=True)

                if totals_mode in ('rows', 'both'):
                    # Итоговая колонка (по строкам) — сумма значений по каждой строке
                    # Находим все value-колонки в pivot
                    val_cols_in_pivot = [c for c in pivot.columns if c not in valid_rows]
                    if val_cols_in_pivot:
                        итого_col_name = 'Итого'
                        pivot[итого_col_name] = pivot[val_cols_in_pivot].sum(axis=1).round(2)

            agg_results.append(pivot)

        if output_format == 'hierarchical':
            # Иерархический формат: каждый (row, agg_function) → отдельная строка
            # Колонки: [row_cols, "__agg__", value_cols_agg1, value_cols_agg2, ...]
            # Для каждой агрегации создаём строки и объединяем вертикально
            hierarchical_rows = []
            row_col_names = [c for c in valid_rows if c in agg_results[0].columns]

            # Собираем все колонки со значениями для каждой агрегации
            all_rows_combined = []
            for idx, agg_func_key in enumerate(agg_functions):
                agg_label = agg_labels.get(agg_func_key, agg_func_key)
                result_df = agg_results[idx].copy()

                # Добавляем колонку AGG_COLUMN_NAME с именем агрегации
                result_df[AGG_COLUMN_NAME] = agg_label

                # Собираем колонки: row_cols + __agg__ + value_cols
                value_cols_in_result = [c for c in result_df.columns
                                        if c not in valid_rows and c != AGG_COLUMN_NAME]
                order_cols = list(valid_rows) + [AGG_COLUMN_NAME] + value_cols_in_result
                result_df = result_df[[c for c in order_cols if c in result_df.columns]]
                all_rows_combined.append(result_df)

            # Объединяем все агрегации
            combined = pd.concat(all_rows_combined, ignore_index=True)

            # Сортируем по row колонкам и AGG_COLUMN_NAME
            sort_cols = list(valid_rows) + [AGG_COLUMN_NAME]
            combined = combined.sort_values(by=sort_cols).reset_index(drop=True)

            # Строки "Итого" должны быть в самом конце, независимо от алфавитной сортировки.
            # Например, "Май" (М) > "Итого" (И) по алфавиту, поэтому после sort_values
            # "Итого" оказывается перед "Май". Перемещаем их в конец.
            итого_mask = combined[AGG_COLUMN_NAME] != combined[AGG_COLUMN_NAME]  # всегда False
            for rc in valid_rows:
                rc_mask = combined[rc].astype(str).str.strip() == 'Итого'
                итого_mask = итого_mask | rc_mask
            if итого_mask.any():
                итого_rows = combined[итого_mask]
                data_rows = combined[~итого_mask]
                combined = pd.concat([data_rows, итого_rows], ignore_index=True)

            # Округление числовых значений
            for col in combined.columns:
                if col not in valid_rows and col != AGG_COLUMN_NAME:
                    try:
                        combined[col] = combined[col].round(2)
                    except Exception as e:
                        logger.warning('Не удалось округлить колонку %s: %s', col, e)

            total_rows = len(combined)
            pivot_data = combined.to_dict(orient='records')
            pivot_columns = combined.columns.tolist()

            return {
                'pivot_data': pivot_data,
                'columns': pivot_columns,
                'total_rows': total_rows,
                'aggregations': agg_functions,
                'output_format': 'hierarchical',
                'row_columns': valid_rows,
                'agg_column': AGG_COLUMN_NAME,
            }
        else:
            # Плоский (широкий) формат: объединяем все агрегации по row колонкам
            row_col_names = [c for c in valid_rows if c in agg_results[0].columns]
            merged = agg_results[0]
            for result_df in agg_results[1:]:
                merge_cols = row_col_names
                merged = pd.merge(merged, result_df, on=merge_cols, how='outer')

            # Заполняем NaN нулями
            merged = merged.fillna(0)

            # Сортируем по row-колонкам, чтобы одинаковые значения шли подряд
            if row_col_names:
                merged = merged.sort_values(by=row_col_names).reset_index(drop=True)

            # Округление числовых значений
            for col in merged.columns:
                if col not in valid_rows:
                    try:
                        merged[col] = merged[col].round(2)
                    except Exception:
                        pass

            total_rows = len(merged)
            pivot_data = merged.to_dict(orient='records')
            pivot_columns = merged.columns.tolist()

            return {
                'pivot_data': pivot_data,
                'columns': pivot_columns,
                'total_rows': total_rows,
                'aggregations': agg_functions,
                'output_format': 'flat',
                'row_columns': row_col_names,
            }

    except Exception as e:
        logger.error('Ошибка построения сводной таблицы:\n%s', traceback.format_exc())
        return {
            'error': f'Ошибка построения сводной таблицы: {str(e)}',
            'pivot_data': [],
            'columns': [],
            'total_rows': 0,
            'aggregations': agg_functions,
            'output_format': output_format,
        }


def save_pivot_to_xlsx(
    pivot_data: List[Dict[str, Any]],
    columns: List[str],
    file_path: str,
    row_columns: Optional[List[str]] = None,
) -> str:
    """
    Сохраняет сводную таблицу в XLSX с объединением одинаковых ячеек
    в row-колонках и выделением строк/столбцов "Итого".

    :param pivot_data: Данные сводной таблицы
    :param columns: Колонки
    :param file_path: Путь для сохранения
    :param row_columns: Список row-колонок для объединения
    :return: file_path
    """
    df = pd.DataFrame(pivot_data, columns=columns)

    # Если нет row_columns — сохраняем как есть
    if not row_columns:
        df.to_excel(file_path, index=False, float_format='%.2f')
        return file_path

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Сводная'

        # --- Заголовки ---
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        itogo_header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            if 'Итого' in col_name:
                cell.fill = itogo_header_fill
            else:
                cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # --- Данные ---
        itogo_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
        itogo_font = Font(bold=True, color='E65100')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        # Определяем индексы row-колонок (0-based)
        row_col_indices = [columns.index(rc) for rc in row_columns if rc in columns]

        # Размечаем объединения: для каждой row-колонки список (start_row, end_row)
        merge_ranges = {}  # col_idx -> [(start, end)]
        for rc_idx in row_col_indices:
            ranges = []
            row = 0
            while row < len(pivot_data):
                val = str(pivot_data[row].get(columns[rc_idx], '') or '')
                # Ищем конец группы одинаковых значений
                end = row
                while end < len(pivot_data) and str(pivot_data[end].get(columns[rc_idx], '') or '') == val:
                    end += 1
                count = end - row  # количество последовательных одинаковых значений
                if count > 1:
                    excel_start = row + 2  # +2: строка 1 — заголовок
                    excel_end = end + 1    # +1: merge_cells включает конечную строку
                    ranges.append((excel_start, excel_end))
                row = end
            merge_ranges[rc_idx] = ranges

        for row_idx, row_data in enumerate(pivot_data):
            excel_row = row_idx + 2  # +2 из-за заголовка (row 1)

            # Проверка на "Итого"
            is_itogo = any(
                str(row_data.get(c, '') or '').strip() == 'Итого'
                for c in columns
            )

            for col_idx, col_name in enumerate(columns):
                val = row_data.get(col_name, '')
                # Форматируем числа
                if isinstance(val, float):
                    val = round(val, 2)

                cell = ws.cell(row=excel_row, column=col_idx + 1, value=val)
                cell.border = thin_border

                if is_itogo:
                    cell.fill = itogo_fill
                    cell.font = itogo_font
                elif col_idx in row_col_indices:
                    # row-колонки — жирный шрифт
                    cell.font = Font(bold=True, size=10)

            # Скрываем строки с rowspan (кроме первой в группе) — не записываем
            # openpyxl не поддерживает rowspan, поэтому для объединения используем merge_cells

        # Объединяем ячейки
        for rc_idx, ranges in merge_ranges.items():
            col_letter = get_column_letter(rc_idx + 1)
            for start_row, end_row in ranges:
                if end_row > start_row:
                    try:
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                        # Выравнивание по центру для объединённых ячеек
                        ws.cell(row=start_row, column=rc_idx + 1).alignment = Alignment(
                            horizontal='center', vertical='center', wrap_text=True
                        )
                    except Exception:
                        pass  # игнорируем ошибки объединения

        # Автоширина колонок
        for col_idx, col_name in enumerate(columns):
            max_len = len(str(col_name))
            for row_idx in range(min(len(pivot_data), 50)):  # первые 50 строк
                val = str(pivot_data[row_idx].get(col_name, '') or '')
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = min(max_len + 3, 40)

        wb.save(file_path)
    except Exception as e:
        logger.error('Ошибка при форматировании XLSX: %s, сохраняю без форматирования', e)
        df.to_excel(file_path, index=False, float_format='%.2f')

    return file_path


# ==================== СЦЕНАРИИ ====================

def save_scenario(name: str, params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Сохраняет сценарий конструктора в JSON-файл.

    :param name: Имя сценария
    :param params: Параметры {columns, filters, pivot_rows, pivot_values, pivot_cols, agg_function}
    :return: {name, created_at, updated_at}
    """
    _ensure_scenarios_dir()
    filename = secure_filename_for_scenario(name) + '.json'
    filepath = os.path.join(SCENARIOS_DIR, filename)

    now = datetime.now().isoformat()
    scenario = {
        'name': name,
        'updated_at': now,
        'params': params,
    }

    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            existing = json.load(f)
        scenario['created_at'] = existing.get('created_at', now)
    else:
        scenario['created_at'] = now

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(scenario, f, ensure_ascii=False, indent=2)

    return scenario


def list_scenarios() -> List[Dict[str, Any]]:
    """
    Возвращает список всех сохранённых сценариев.

    :return: [{name, created_at, updated_at}, ...]
    """
    _ensure_scenarios_dir()
    scenarios = []
    if not os.path.exists(SCENARIOS_DIR):
        return scenarios

    for fname in sorted(os.listdir(SCENARIOS_DIR)):
        if fname.endswith('.json'):
            try:
                with open(os.path.join(SCENARIOS_DIR, fname), 'r', encoding='utf-8') as f:
                    data = json.load(f)
                scenarios.append({
                    'name': data.get('name', fname[:-5]),
                    'created_at': data.get('created_at', ''),
                    'updated_at': data.get('updated_at', ''),
                })
            except Exception as e:
                logger.warning('Не удалось прочитать сценарий %s: %s', fname, e)
    return scenarios


def load_scenario(name: str) -> Optional[Dict[str, Any]]:
    """
    Загружает сценарий по имени.

    :param name: Имя сценария
    :return: {name, created_at, updated_at, params} или None
    """
    _ensure_scenarios_dir()
    filename = secure_filename_for_scenario(name) + '.json'
    filepath = os.path.join(SCENARIOS_DIR, filename)

    if not os.path.exists(filepath):
        return None

    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def delete_scenario(name: str) -> bool:
    """
    Удаляет сценарий.

    :param name: Имя сценария
    :return: True если удалён, False если не найден
    """
    _ensure_scenarios_dir()
    filename = secure_filename_for_scenario(name) + '.json'
    filepath = os.path.join(SCENARIOS_DIR, filename)

    if not os.path.exists(filepath):
        return False

    os.remove(filepath)
    return True


def secure_filename_for_scenario(name: str) -> str:
    """Приводит имя сценария к безопасному имени файла."""
    import re
    safe = re.sub(r'[^\w\-_\. ]', '_', name)
    return safe.strip().replace(' ', '_')
