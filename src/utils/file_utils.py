# -*- coding: utf-8 -*-
"""
Утилиты для работы с файлами: безопасные имена, очистка загрузок, проверка расширений,
централизованное чтение CSV и Excel.
"""

import os
import re
import time
import uuid
import logging
from typing import Set, Dict, Any, List

logger = logging.getLogger(__name__)


def safe_filename(name: str) -> str:
    """
    Заменяет недопустимые символы на '_', сохраняя кириллицу,
    латиницу, пробелы, точки и дефисы.
    """
    safe = re.sub(r'[^\w\s\u0400-\u04FF.-]', '_', name, flags=re.UNICODE)
    return safe.replace(' ', '_')


def clean_old_uploads(upload_dir: str, max_age_seconds: int = 3600) -> None:
    """
    Удаляет файлы из upload_dir, возраст которых превышает max_age_seconds.
    """
    now = time.time()
    if not os.path.exists(upload_dir):
        return
    for filename in os.listdir(upload_dir):
        filepath = os.path.join(upload_dir, filename)
        try:
            if os.path.isfile(filepath) and (now - os.path.getmtime(filepath)) > max_age_seconds:
                os.remove(filepath)
        except Exception as e:
            logger.warning('Не удалось удалить старый файл %s: %s', filename, e)


ALLOWED_EXTENSIONS: Set[str] = {'txt', 'md'}


def allowed_file(filename: str) -> bool:
    """Проверяет, допустимо ли расширение файла для загрузки в чат."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_text_from_txt(filepath: str) -> str:
    """Пытается прочитать текстовый файл в разных кодировках."""
    encodings = ['utf-8', 'cp1251', 'latin-1']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    return "Не удалось прочитать файл: неизвестная кодировка."


def safe_remove(filepath: str) -> None:
    """Безопасно удаляет файл, игнорируя ошибки."""
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
    except Exception as e:
        logger.warning('Не удалось удалить файл %s: %s', filepath, e)


def generate_temp_path(upload_dir: str, prefix: str = "temp_", suffix: str = ".xlsx") -> str:
    """Генерирует уникальный временный путь в папке upload_dir."""
    return os.path.join(upload_dir, f"{prefix}{uuid.uuid4().hex}{suffix}")


def save_json_as_xlsx(data: List[Dict[str, Any]], filepath: str) -> str:
    """
    Сохраняет JSON-массив объектов как .xlsx файл с форматированием.

    :param data: Список словарей (колонка -> значение)
    :param filepath: Путь для сохранения
    :return: filepath
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    if data:
        headers = list(data[0].keys())
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_idx in range(2, len(data) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(filepath)
    return filepath


def read_file_to_df(file_path: str, sheet_name=None, **kwargs) -> "pd.DataFrame":
    """
    Централизованное чтение CSV (.csv) или Excel (.xlsx/.xls) в pandas DataFrame.

    - Для CSV: автоматически определяет кодировку (utf-8 → cp1251).
      Параметр sheet_name игнорируется (CSV не имеет листов).
    - Для Excel: передаёт sheet_name в pd.read_excel().

    :param file_path: Путь к файлу
    :param sheet_name: Имя листа (только для Excel). По умолчанию 0 (первый лист).
    :param kwargs: Дополнительные параметры, передаваемые в pd.read_csv() / pd.read_excel()
    :return: pandas DataFrame
    """
    import pandas as pd

    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        # Для CSV sheet_name не нужен, удаляем если передан
        kwargs.pop('sheet_name', None)
        try:
            return pd.read_csv(file_path, encoding='utf-8', **kwargs)
        except UnicodeDecodeError:
            return pd.read_csv(file_path, encoding='cp1251', **kwargs)
    else:
        # Excel: .xlsx, .xls
        return pd.read_excel(file_path, sheet_name=sheet_name or 0, **kwargs)
